import random
import openpyxl
import math

# ハウスルーツ用の追加ダメージ計算機

def getadditionaldamage(dice: int, hantei: str) -> int:
    if hantei == 'miss':
        None

    elif hantei == 'normal':
        dicedata = str(dice)
        dicelist = (list(dicedata))
        return int(dicelist[0])

    elif hantei == 'hard':
        dicedata = str(dice)
        dicelist = (list(dicedata))
        if len(dicelist) == 2:
            return int(dicelist[0])*2
        elif len(dicelist) == 1:
            return 0

    elif hantei == 'ex':
        dicedata = str(dice)
        dicelist = (list(dicedata))
        if len(dicelist) == 2:
            return int(dicelist[0])
        elif len(dicelist) == 1:
            return 0

# エクセルからデータを取ってくる

class Getcharacterdata:
    def __init__(self, xlsx: str, column: str):
        self.wb = openpyxl.load_workbook(xlsx)
        self.column = column
        self.sheet = self.wb['探索者情報']
        self.name = self.sheet['A'+self.column]
        self.skillvalue = self.sheet['F'+self.column]
        self.weapondmg = None
        self.weapondmg1 = self.sheet['B'+self.column]
        self.weapondmg2 = self.sheet['C'+self.column]
        self.weapondmg3 = self.sheet['D'+self.column]
        self.dmgbonus = self.sheet['E'+self.column]
        self.dbonoff = self.sheet['G'+self.column]

    def printtest(self):
        print(self.name.value, self.skillvalue.value, self.weapondmg.value, self.dmgbonus.value)

    def namereturn(self) -> str:
        return self.name.value

    def skillvaluereturn(self) -> str:
        return self.skillvalue.value

    def weapondmgreturn(self) -> float:
        if self.weapondmg1.value is None:
            self.weapondmg4 = 0

        elif 'd' in self.weapondmg1.value:
            self.weapondmg4 = str(Dice(self.weapondmg1.value).replacediceroll())

        else:
            self.weapondmg4 = self.weapondmg1.value

        if self.weapondmg2.value is None:
            self.weapondmg5 = 0

        elif 'd' in self.weapondmg2.value:
            self.weapondmg5 = str(Dice(self.weapondmg2.value).replacediceroll())

        else:
            self.weapondmg5 = self.weapondmg2.value

        if self.weapondmg3.value is None:
            self.weapondmg6 = 0

        elif 'd' in self.weapondmg3.value:
            self.weapondmg6 = str(Dice(self.weapondmg3.value).replacediceroll())

        else:
            self.weapondmg6 = self.weapondmg3.value

        self.weapondmg = float(self.weapondmg4) + float(self.weapondmg5) + float(self.weapondmg6)
        return self.weapondmg

    def dmgbonusreturn(self) -> str:
        if self.dbonoff.value == 'on':
            if self.dmgbonus.value == '1d4':
                return Dice('1d4').replacediceroll()
            elif self.dmgbonus.value == '1d6':
                return Dice('1d6').replacediceroll()
            else:
                return self.dmgbonus.value

        elif self.dbonoff.value == 'off':
            return 0

        elif self.dbonoff.value == '1/2':
            if self.dmgbonus.value == '1d4' or '1D4' :
                return Dice('1d4').replacediceroll()/2
            else:
                return self.dmgbonus.value/2

# 技能のハード判定、ＥＸ判定の数値算出用

def hardexdata(ginouvalue: str, situation: str) -> str:
    if situation == 'hard':
        return math.floor(ginouvalue/2)
    elif situation == 'ex':
        return math.floor(ginouvalue/5)

# エクセルの各データと追加ダメージ計算機で出た結果を元にダメージの計算を行う

def dmgcalculation(ginoudice: int, column: str) -> str:

    ginoudice = ginoudice
    ginou = int(Getcharacterdata('data.xlsx', column).skillvaluereturn())
    ginouhard = hardexdata(ginou, 'hard')
    ginouex = hardexdata(ginou, 'ex')
    hantei = None

    if ginoudice>=90:
        hantei = 'nohit'

        return 'miss'

    elif ginoudice>ginou:
        hantei = 'miss'

        return '0'

    elif ginoudice>ginouhard:
        hantei = 'normal'
        additionaldamage = str(getadditionaldamage(ginoudice, hantei))

        return additionaldamage

    elif ginoudice>ginouex:
        hantei = 'hard'
        additionaldamage = str(getadditionaldamage(ginoudice, hantei))

        return additionaldamage

    elif ginoudice<=ginouex:
        hantei = 'ex'
        additionaldamage = float(getadditionaldamage(ginoudice, hantei))+float(11)
        return additionaldamage

# 〇ｄ〇みたいなダイスロールの表記を期待値に置き換える

class Dice:
    def __init__(self, dice_notation: str):
        dice_notation.replace('D', 'd')
        dice_data_list = dice_notation.split('d')

        self.times = int(dice_data_list[0])
        self.faces = int(dice_data_list[1])

    def roll(self) -> int:
        result = 0
        for _ in range(self.times):
            result += random.randint(1, self.faces)
        return result

    def replacediceroll(self) -> int:
        result = 0
        for i in range(self.faces):
            result += (i + 1) / self.faces
        result = result * self.times
        return result

# １～１００までのダメージを計算して期待値を求める

def expectedvalue(xlsx: str, column: str):
    result = 0
    for i in range(100):
        additionaldmg = dmgcalculation(i+1, column)
        weapondmg = Getcharacterdata(xlsx, column).weapondmgreturn()
        dmgbonus = float(Getcharacterdata(xlsx, column).dmgbonusreturn())
        if additionaldmg == 'miss':
            result += 0
        else:
            additionaldmg = float(additionaldmg)
            result += (additionaldmg + weapondmg + dmgbonus)
    return result / 100

# 計算結果をエクセルに書き込む

def printxlsx(xlsx: str, expectedvaluedata: float, column: str):
    wb = openpyxl.load_workbook(xlsx)
    sheet = wb['探索者情報']
    sheet['I' + column].value = expectedvaluedata
    wb.save("data.xlsx")

# エクセルを読み込んでキャラクターデータがある場所のみ読み込ませる

def charactercheck(xlsx: str):
    wb = openpyxl.load_workbook(xlsx)
    sheet = wb['探索者情報']
    columnlist = []
    for i in range(10):
        if sheet['A' + str(i+2)].value is None:
            None
        else:
            columnlist.append(i+2)
    return columnlist

# パーティ全体の期待ダメージ量を計算する

def getpartyexpectedvalue(xlsx: str):
    wb = openpyxl.load_workbook(xlsx)
    sheet = wb['探索者情報']
    partyexpectedvalue = 0
    for i in range(10):
        if sheet['I' + str(i+2)].value is None:
            None
        else:
            partyexpectedvalue += sheet['I' + str(i+2)].value
    sheet['J2'].value = partyexpectedvalue
    wb.save("data.xlsx")

class Getenemydata:
    def __init__(self, xlsx: str):
        self.wb = openpyxl.load_workbook(xlsx)
        self.column = '2'
        self.sheet = self.wb['敵情報']
        self.name = self.sheet['A'+self.column]
        self.defensepower = self.sheet['C'+self.column]

    def printtest(self):
        print(self.name.value, self.defensepower.value)

    def namereturn(self) -> str:
        return self.name.value

    def enemydefensepowerreturn(self) -> str:
        if self.defensepower.value is None:
            return 0
        else:
            return self.defensepower.value

def fullexpectedvalue(xlsx: str, column: str):
    result = 0
    for i in range(100):
        additionaldmg = dmgcalculation(i+1, column)
        weapondmg = Getcharacterdata(xlsx, column).weapondmgreturn()
        dmgbonus = float(Getcharacterdata(xlsx, column).dmgbonusreturn())
        enemydefensepower = Getenemydata(xlsx).enemydefensepowerreturn()
        if additionaldmg == 'miss':
            result += 0
        else:
            additionaldmg = float(additionaldmg)
            result += ((additionaldmg + weapondmg + dmgbonus) - float(enemydefensepower))
    return result / 100

def fullprintxlsx(xlsx: str, expectedvaluedata: float, column: str):
    wb = openpyxl.load_workbook(xlsx)
    sheet = wb['探索者情報']
    sheet['K' + column].value = expectedvaluedata
    wb.save("data.xlsx")

def getfullexpectedvalue(xlsx: str):
    wb = openpyxl.load_workbook(xlsx)
    sheet = wb['探索者情報']
    partyexpectedvalue = 0
    for i in range(10):
        if sheet['K' + str(i+2)].value is None:
            None
        else:
            partyexpectedvalue += sheet['K' + str(i+2)].value
    sheet['L2'].value = partyexpectedvalue
    wb.save("data.xlsx")

# FEOダメージの期待値生成器

class Feoexpectedvalue:
    def __init__(self, xlsx:str):
        self.wb = openpyxl.load_workbook(xlsx)
        self.sheet = self.wb['FEO期待値表']

    def feodmgprint(self):
        for i in range(20):
            self.sheet['B' + str(i+2)].value = '1d' + str(i+1) + '+3'
            self.wb.save("data.xlsx")


