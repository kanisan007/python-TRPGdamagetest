import openpyxl

# openpyxlを使ったエクセルファイルの読み込みテスト

# wb = openpyxl.load_workbook('data.xlsx')
# print(wb.sheetnames)

# sheet = wb['kiekie']
# cell = sheet['D2']

# print(cell.value)

# def getcelldata(xlsx: str, column: str, value: str) -> str:
#     wb = openpyxl.load_workbook(xlsx)
#     sheet = wb['kiekie']
#     cell = sheet['D2']
#     return cell

# kie = getcelldata('data.xlsx', '2', 'D')
# print(kie)

class Getcharacterdata:
    def __init__(self, xlsx: str, column: str):
        self.wb = openpyxl.load_workbook(xlsx)
        self.column = column
        self.sheet = self.wb['kiekie']
        self.name = self.sheet['A'+self.column]
        self.skillvalue = self.sheet['D'+self.column]
        self.weapondmg = self.sheet['B'+self.column]
        self.dmgbonus = self.sheet['C'+self.column]

    def printkie(self):
        print(self.name.value, self.skillvalue.value, self.weapondmg.value, self.dmgbonus.value)

    def namereturn(self):
        return self.name.value

    def skillvaluereturn(self):
        return self.skillvalue.value

    def weapondmgreturn(self):
        return self.dmgbonus.value

    def dmgbonusreturn(self):
        return self.dmgbonus.value

a = Getcharacterdata('data.xlsx', '2').printing()




